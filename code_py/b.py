import sys
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication, QFileDialog, QMainWindow
from PyQt5.uic import loadUi
from PyQt5.QtCore import QIODevice, QObject, QThread, pyqtSignal
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo
from openpyxl import Workbook, load_workbook

class MainWindow(QMainWindow,QDialog):
    def __init__(self):
        super(MainWindow,self).__init__()
        loadUi("test.ui",self)
        self.data = [0,0]
        self.dc1 = 0
        self.dc2 = 0
        self.dc3 = 0
        self.dc4 = 0
        self.sig = 0
        self.serial = QSerialPort()
        self.serial.setBaudRate(115200)
        self.browse.clicked.connect(self.browsefiles)
        self.openSerial.clicked.connect(self.openSer)
        self.closeSerial.clicked.connect(self.closeSer)
        self.Save.clicked.connect(self.save_state)
        self.Run.clicked.connect(self.run)
        self.ManualMode_.clicked.connect(self.ManualMode)
        self.Automode_.clicked.connect(self.AutoMode)
        self.list_state.addItem("1")
        self.list_state.addItem("2")
        self.list_state.addItem("3")
        self.list_state.addItem("4")
        self.list_state.addItem("5")
        self.list_state.addItem("6")
        self.list_state.addItem("7")
        self.list_state.addItem("8")
        self.list_state.addItem("9")
        self.list_state.addItem("10")

    def run(self):
        wb = load_workbook(filename=self.filename.text())
        sheet = wb['Sheet1']
        i = 0
        while True:
            if(self.sig == 0):
                pub_dc1 = sheet['B'+str(i+2)].value
                pub_dc2 = sheet['C'+str(i+2)].value
                pub_dc3 = sheet['D'+str(i+2)].value
                pub_dc4 = sheet['E'+str(i+2)].value
                a = [pub_dc1,pub_dc2,pub_dc3,pub_dc4]
                self.serialSend(a)
                # print("State thu ",i)
                i = i + 1
                self.sig = 1

            if(i == 10):
                break
                
    def serialSend(self,data):
        txs = ""
        for val in data:
            txs += str(val)
            txs += ','    
        txs = txs.rstrip(',')
        txs  += ';'
        self.serial.write(txs.encode())

    def save_state(self):
        wb = load_workbook(filename=self.filename.text())
        sheet = wb['Sheet1']
        state = int(self.list_state.currentText())
        sheet['B'+str(state+1)].value = int(self.showDC1.text())
        sheet['C'+str(state+1)].value = int(self.showDC2.text())
        sheet['D'+str(state+1)].value = int(self.showDC3.text())
        sheet['E'+str(state+1)].value = int(self.showDC4.text())
        wb.save(filename=self.filename.text())

    def get_Serial(self):
        port_list = []
        ports = QSerialPortInfo().availablePorts()
        for port in ports:
            port_list.append(port.portName())
        self.comlist.addItems(port_list)

    def onRead(self):
        rx = self.serial.readLine()
        cvt_rx = str(rx, 'utf-8').strip()
        self.data = cvt_rx.split(',')
        print(self.data)
        if(len(self.data) == 4):
            self.dc1 = self.data[0]
            self.dc2 = self.data[1]
            self.dc3 = self.data[2]
            self.dc4 = self.data[3]
            self.showDC1.setText(str(self.dc1))
            self.showDC2.setText(str(self.dc2))
            self.showDC3.setText(str(self.dc3))
            self.showDC4.setText(str(self.dc4))
        if(self.data == "D"):
            self.sig = 0
             
    def ManualMode(self):
        send_var = "M"
        self.serial.write(send_var.encode())
    
    def AutoMode(self):
        send_var = "A"
        self.serial.write(send_var.encode())

    def browsefiles(self):
        fname=QFileDialog.getOpenFileName(self, 'Open file', '/home/tucuman/Arduino/Project/code_py')
        self.filename.setText(fname[0])
    
    def openSer(self):
        self.serial.setPortName(self.comlist.currentText())
        self.serial.open(QIODevice.ReadWrite)
        print("Da ket noi voi Arduino")
        self.serial.readyRead.connect(self.onRead)

    def closeSer(self):
        self.serial.close()
        print("Da ngat ket noi")

if __name__ == '__main__':
    app=QApplication(sys.argv)
    mainwindow=MainWindow()
    mainwindow.get_Serial()
    mainwindow.onRead()
    widget=QtWidgets.QStackedWidget()
    widget.addWidget(mainwindow)
    widget.setFixedWidth(640)
    widget.setFixedHeight(430)
    widget.setWindowTitle("Control Panel")
    widget.show()
    sys.exit(app.exec_())